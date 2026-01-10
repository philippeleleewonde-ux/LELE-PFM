#!/usr/bin/env python3
"""
Script d'extraction exhaustive de la structure de la feuille 13-REPORTING M1-Pdf
pour créer le plan de développement de la Page 17: Global Reporting HCM Performance Plan
"""

import openpyxl
from collections import defaultdict
import json

# Chemin vers le fichier Excel
EXCEL_PATH = "/Users/onclephilbasket/Documents/Sauvergarde docs Macbookair15/Projet Modules HCM ACCOUNTING/MODULE 1/Sources Excel/2-M1-Module de calcul-test2ok200525.xlsx"

def extract_reporting_structure():
    """Extraire la structure complète de la feuille 13-REPORTING M1-Pdf"""

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    ws = wb['13-REPORTING M1-Pdf']

    print("=" * 120)
    print("EXTRACTION EXHAUSTIVE - FEUILLE 13-REPORTING M1-Pdf")
    print("=" * 120)
    print()

    # Structure pour stocker les données
    structure = {
        'sections': [],
        'formulas_by_source': defaultdict(list),
        'all_formulas': [],
        'total_rows': ws.max_row,
        'total_cols': ws.max_column
    }

    print(f"📊 Dimensions feuille: {ws.max_row} lignes × {ws.max_column} colonnes")
    print()

    # Extraire toutes les formules et leur source
    print("🔍 Extraction de toutes les formules...")
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                formula_data = {
                    'coord': cell.coordinate,
                    'row': cell.row,
                    'col': cell.column,
                    'formula': cell.value
                }

                structure['all_formulas'].append(formula_data)

                # Identifier la feuille source
                if '!' in cell.value:
                    sheet_ref = cell.value.split('!')[0].replace('=', '').replace("'", "")
                    structure['formulas_by_source'][sheet_ref].append(formula_data)

    print(f"✅ {len(structure['all_formulas'])} formules extraites")
    print()

    # Grouper par feuille source
    print("📑 Formules groupées par feuille source:")
    for sheet_name, formulas in sorted(structure['formulas_by_source'].items(), key=lambda x: len(x[1]), reverse=True):
        print(f"  - {sheet_name:50s}: {len(formulas):4d} formules")
    print()

    # Identifier les sections principales (lignes 1-200)
    print("🎯 Identification des sections principales (lignes 1-200):")
    for row_num in range(1, 201):
        row = ws[row_num]
        cell_a = row[0].value  # Colonne A
        cell_b = row[1].value if len(row) > 1 else None  # Colonne B

        # Chercher les titres importants
        if cell_a and isinstance(cell_a, str) and len(cell_a) > 15:
            section_type = None

            # Identifier le type de section
            if any(keyword in cell_a.lower() for keyword in ['value at risk', 'var', 'unexpected']):
                section_type = 'A - Value at Risk'
            elif any(keyword in cell_a.lower() for keyword in ['distribution of costs', 'savings', 'indicators']):
                section_type = 'B - Distribution Savings'
            elif any(keyword in cell_a.lower() for keyword in ['basel', 'operational risk', 'loss events']):
                section_type = 'C - Distribution VaR Basel'
            elif any(keyword in cell_a.lower() for keyword in ['threshold', 'tolerance', 'historic']):
                section_type = 'D - Historic Threshold'
            elif any(keyword in cell_a.lower() for keyword in ['3 year', 'cash flow', 'budget']):
                section_type = 'E - 3 Year Budget'
            elif any(keyword in cell_a.lower() for keyword in ['scr', 'impact']):
                section_type = 'F - SCR Impact'

            if section_type:
                structure['sections'].append({
                    'type': section_type,
                    'row': row_num,
                    'title': cell_a
                })
                print(f"  Ligne {row_num:3d}: {section_type:30s} | {cell_a[:60]}")

    print()
    print(f"✅ {len(structure['sections'])} sections identifiées")
    print()

    # Exporter en JSON pour analyse
    output_file = "docs/REPORTING_STRUCTURE_ANALYSIS.json"
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(structure, f, indent=2, ensure_ascii=False)

    print(f"💾 Structure exportée vers: {output_file}")
    print()

    wb.close()

    return structure

if __name__ == "__main__":
    structure = extract_reporting_structure()

    print("=" * 120)
    print("RÉSUMÉ FINAL")
    print("=" * 120)
    print(f"Total formules: {len(structure['all_formulas'])}")
    print(f"Feuilles sources: {len(structure['formulas_by_source'])}")
    print(f"Sections identifiées: {len(structure['sections'])}")
    print()

    # Mapper vers les pages React
    print("📌 MAPPING VERS PAGES REACT:")
    sheet_to_page = {
        '5-CALCUL VALEUR EN RISQUE': 'Page 7 (PRL Accounts)',
        '6-ECONOMIES SUR 3 ANS': 'Pages 8-9 (EE/IPLE)',
        '7-REPORTING COUTS INCIDENTS': 'Page 10 (Economic Breakdown)',
        '8-SEUIL HISTORIQUE-SURCOUTS': 'Page 11 (Risk Threshold)',
        '9-PLANIFICATION BUDGET-3ANS': 'Page 12-13 (3-Year Plan/Dashboard)'
    }

    for sheet, page in sheet_to_page.items():
        count = len(structure['formulas_by_source'].get(sheet, []))
        print(f"  {sheet:40s} → {page:40s} ({count} formules)")
